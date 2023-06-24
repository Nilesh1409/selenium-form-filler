from selenium import webdriver
import selenium
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, UnexpectedAlertPresentException
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import UnexpectedAlertPresentException
import pyperclip
import openpyxl
import time
import datetime

def format_date(dt):
    if isinstance(dt, datetime.datetime):
        return dt.strftime("%d/%m/%Y")
    else:
        return dt

# Load the Excel file
print("Started")
workbook = openpyxl.load_workbook('./form-data.xlsx')
worksheet = workbook.active

# Specify the path to the Chrome User Data Directory
chrome_options = Options()
chrome_options.add_argument("--user-data-dir=/Users/nilesh/Library/Application Support/Google/Chrome")
print("after getting user dir")
# Set the path to the Chrome driver executable
chrome_driver_path = "/Users/nilesh/Downloads/chromedriver_mac_arm64/chromedriver"

# Specify the path to the Chrome binary (replace with your actual Chrome binary path)
chrome_options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"

# Start the Chrome webdriver with the specified options
driver = webdriver.Chrome(service=Service(chrome_driver_path), options=chrome_options)

# Open the URL
print("Started")
index = 2
driver.get('https://rskilling.reliancefoundation.org/candidate/create')
row_number = 120 + index

# Iterate over the rows in the Excel file
for row in worksheet.iter_rows(min_row=108, values_only=True):
    index+1
    # Extract data from Excel
    split_gmail = row[16].split('@')
    name = row[8]
    email = split_gmail[0] +'@gmail.com'
    phone = row[17]
    gander = row[12]
    dob = row[4]
    doj = row[14]
    pan = row[40]
    pyperclip.copy(name)
    print(row)

    try:
        print("in form block")
        time.sleep(3)
        # Find the form elements and fill them
        driver.find_element(By.XPATH, '//*[@id="cand_form"]/div[1]/div/div[1]/div[1]/span/span[1]/span/span[2]').click()
        driver.find_element(By.XPATH, '/html/body/span/span/span[2]/ul/li[1]').click()
        time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[1]/div[2]/div[1]/span/span[1]/span/span[2]').click()
        driver.find_element(By.XPATH, '/html/body/span/span/span[2]/ul/li[2]').click()
        # time.sleep(5)
        nameInput = driver.find_element(By.XPATH,'/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[2]/div[1]/div/input')
        nameInput.send_keys(name)
        # time.sleep(2)
        driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[2]/div[2]/div/span/span[1]/span/span[2]').click()
        if gander=='M':
            driver.find_element(By.XPATH, '/html/body/span/span/span[2]/ul/li[1]').click()
        else: 
            driver.find_element(By.XPATH,'/html/body/span/span/span[2]/ul/li[2]').click()
        # time.sleep(2)
        dobInput = driver.find_element(By.XPATH,'/html/body/div[1]/div[1]/section[2]/div/div/form/div[1]/div/div[2]/div[3]/div/div[1]/input')
        dobInput.send_keys(format_date(dob))
        # time.sleep(2)
        dobInput = driver.find_element(By.XPATH,'/html/body/div[1]/div[1]/section[2]/div/div/form/div[1]/div/div[3]/div[1]/div/input')
        dobInput.send_keys(format_date(doj))
        # time.sleep(1)
        emailInput = driver.find_element(By.XPATH,'/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[3]/div[2]/div/input')
        
        emailInput.send_keys(email)
        # time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[4]/div/span/span[1]/span/span[2]').click()
        driver.find_element(By.XPATH, '/html/body/span/span/span[2]/ul/li[2]').click()
        # time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[6]/div/select').click()
        driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[6]/div/select/option[2]').click()
        # time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[7]/div/select').click()
        driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[7]/div/select/option[2]').click()
        time.sleep(1)
        panInput = driver.find_element(By.XPATH,'/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[8]/div/div/input')
        panInput.send_keys(pan)
        time.sleep(1)
        phoneInput = driver.find_element(By.XPATH,'/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[10]/div/input')
        phoneInput.send_keys(phone)
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[11]/div[1]/div/span/span[1]/span/span[2]').click()
        stateInput = driver.find_element(By.XPATH,'/html/body/span/span/span[1]/input')
        stateInput.send_keys('Mah')
        # time.sleep(1)
        stateInput.send_keys(Keys.RETURN)
        # driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[11]/div[2]/div/select').click()
        time.sleep(1)
        # distInput = driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[11]/div[2]/div/select').click()
        # distInput.send_key("mum")
        # time.sleep(3)
        # dist = driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[11]/div[2]/div/select/option[18]')
        # print("distttt",dist)
        # dist.click
        # Locate the select element
        distSelect = driver.find_element(By.ID, 'depDropDistrict')

# Create a Select object with the select element
        select = Select(distSelect)

# Select the option by value
        time.sleep(1)
        select.select_by_value('394')
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[11]/div[3]/div/select').click()
        time.sleep(1)
        driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[1]/div/div[11]/div[3]/div/select/option[7]').click()
        time.sleep(1)

        driver.find_element(By.XPATH, '/html/body/div/div[1]/section[2]/div/div/form/div[2]/h3[2]').click()

        
        time.sleep(15)
        # select_box1 = Select(driver.find_element(By.XPATH, '//*[@id="center-id"]'))
        # select_box1 = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="center-id"]')))
        # select_box1.select_by_visible_text('Pro Consultant Mumbai')  # Index starts from 0, so 6 represents the 7th option
        print("in form block 2")
        
        time.sleep(3)  # Wait for 3 seconds
        photoInput = driver.find_element(By.XPATH,'/html/body/div/div[1]/section[2]/div/div/form/div[2]/div[2]/div[8]/input')
        uploaded_file = photoInput.get_attribute('value')
        skip = False
        try:
            while not uploaded_file:
                photoInput = driver.find_element(By.XPATH,'/html/body/div/div[1]/section[2]/div/div/form/div[2]/div[2]/div[8]/input')
                uploaded_file = photoInput.get_attribute('value')
                try:
                    skip_file = driver.find_element(By.XPATH,'/html/body/div[1]/div[1]/section[2]/div/div/form/div[2]/div[2]/div[9]/input')
                    skip_file_value = skip_file.get_attribute('value')
                    print("skip_file_value",skip_file_value, skip)
                    if skip_file_value:
                        skip = True
                        uploaded_file = True
                    print('skip',skip)
                    if skip:
                        worksheet.cell(row=row_number, column=1).value = str(row)
                        driver.refresh()
                        continue
                except UnexpectedAlertPresentException:
                    driver.switch_to.alert.dismiss()
       

            print('file uploaded')
            time.sleep(2)
            driver.find_element(By.XPATH,'/html/body/div/div[1]/section[2]/div/div/form/div[3]/button').click()
            time.sleep(5)
        
        # time.sleep(30)  # Wait for 3 seconds
        
        # Select the 3rd option in the second select box
        # select_box2 = Select(driver.find_element(By.XPATH, '//*[@id="candidate-batch_id"]'))
        # select_box2.select_by_visible_text('MH-MUMB-MUMB-RET-PRO-22')  # Index starts from 0, so 2 represents the 3rd option
            print("in form block 3")
        except UnexpectedAlertPresentException:
        # Handle the unexpected alert here
        # For example, you can dismiss the alert using:
            driver.switch_to.alert.dismiss()
        
    except NoSuchElementException:
        # If name_field is not found, perform the necessary actions
        email = driver.find_element(By.XPATH, '//*[@id="loginform-username"]')
        email.send_keys('proconsultant.hr@gmail.com')
        time.sleep(1)
        password = driver.find_element(By.XPATH, '//*[@id="myPassword"]')
        password.send_keys('Rfskilling@123')
        login = driver.find_element(By.XPATH, '//*[@id="login-form"]/div[4]/div/button')
        print("login",login)
        login.click()
        time.sleep(15)
        driver.get('https://rskilling.reliancefoundation.org/candidate/create')
        continue
    print("after login")
    

#    108

    # Wait for the page to load (optional)
    verify_upload = driver.find_element(By.XPATH,'/html/body/div[1]/div[1]/section[2]/p/a')
    while not verify_upload:
        verify_upload = driver.find_element(By.XPATH,'/html/body/div[1]/div[1]/section[2]/p/a')
        time.sleep(2)
    print(verify_upload,'verify_upload')
    # time.sleep(60)
    driver.get('https://rskilling.reliancefoundation.org/candidate/create')

    driver.implicitly_wait(5)

# Close the browser
driver.quit()
